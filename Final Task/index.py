from worker import *
from actions import *
import openpyxl

#Відкриття файлу з перехопленням помилки
f = True
while f:
    try:
        filename = input('Enter file name: ')  # workers.xlsx
        wb = openpyxl.load_workbook(filename)
        break
    except FileNotFoundError:
        print('File does not exist or format is incorrect')

#Активний лист
sheet = wb.get_sheet_by_name('wlist')

#Заповнення двовимірного робочого масиву:
data1 = []
l = []
rows = 2 # Початок зчитування з 2-го рядка, та як 1 - Title
a = sheet.cell(row=rows, column=1).value # Значення з 1 комірки рядка (id)

while a:  # Поки комірка заповнена
    for i in range(1,6):
        l.append(sheet.cell(row=rows, column=i).value) # Створюється масив з даними для конкретного працівника
    data1.append(l) # Додавання даних в основний робчий масив
    l = []
    rows += 1
    a = sheet.cell(row=rows, column=1).value

#Дані, витягнуті з файлу
print('\n', 'Data from file:')
for i in data1:
    print(i)

#Створення списку працівників W (об'єктів Worker)
w = []
for i in range(0,len(data1)):
    if data1[i][4] == 'F':
        a = Fixed(int(data1[i][0]),data1[i][1],data1[i][2],int(data1[i][3]))
        w.append(a)
    else:
        a = PerHour(int(data1[i][0]), data1[i][1], data1[i][2], int(data1[i][3]))
        w.append(a)

print('\n', 'List of workers:')
for i in w:
    i.out()

#Сортування працівників
w_sorted = sorting(w)

print('\n', 'Sorted list of workers:')
for i in w_sorted:
    i.out()

print('\n', 'First 5 workers:')
firstfive(w_sorted)

print('\n', 'Last 3 workers:')
lastthree(w_sorted)

#Створення нового файлу
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

#Копіювання заголовків стовпців з 1 таблиці (тобто копія 1-го рядка)
for i in range(1,5):
   sheet2.cell(row=1,column=i).value = sheet.cell(row=1, column=i).value

#Запис даних в новий файл
r = 2
for i in w_sorted:
    c = 1
    while c < 5:
        if c == 1:
            sheet2.cell(row=r,column=c).value = i.id
        elif c == 2:
            sheet2.cell(row=r, column=c).value = i.name
        elif c == 3:
            sheet2.cell(row=r, column=c).value = i.surname
        elif c == 4:
            sheet2.cell(row=r, column=c).value = i.salary()
        c += 1
    r += 1

#Збереження файлу
wb2.save('Sorted.xlsx')